#!/usr/bin/env python3
"""Build a deterministic, grade-stratified GAMMA glaucoma manifest."""
from __future__ import annotations
import argparse, csv, json, random
from collections import Counter, defaultdict
from pathlib import Path

GRADES = ("non", "early", "mid_advanced")

def arguments():
    root = Path(__file__).resolve().parents[1]
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--raw-dir", type=Path, default=root / "data_gamma/raw")
    p.add_argument("--manifest", type=Path, default=root / "data_gamma/manifest.csv")
    p.add_argument("--summary", type=Path, default=root / "data_gamma/prepare_summary.json")
    p.add_argument("--seed", type=int, default=2026)
    return p.parse_args()

def allocate(ids, rng):
    ids = sorted(ids); rng.shuffle(ids)
    train, val = round(len(ids) * .6), round(len(ids) * .2)
    return {x: "train" if i < train else "val" if i < train + val else "test" for i, x in enumerate(ids)}

def rel(path, root):
    # Keep the repository-facing path when `data_gamma/raw` is a symlink into
    # cluster storage. Resolving first would turn it into `/data/...` and make
    # the otherwise valid path appear to be outside OphthalmicAgent.
    path = path.expanduser().absolute()
    root = root.expanduser().absolute()
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path.resolve())

def main():
    args = arguments()
    from openpyxl import load_workbook
    root = Path(__file__).resolve().parents[1]
    grading = args.raw_dir / "grading/Glaucoma_grading/training"
    labels = grading / "glaucoma_grading_training_GT.xlsx"
    modalities = grading / "multi-modality_images"
    sheet = load_workbook(labels, read_only=True, data_only=True).active
    source, strata = [], defaultdict(list)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        case = str(values[0]).zfill(4); one_hot = [int(x) for x in values[1:4]]
        if sum(one_hot) != 1: raise SystemExit(f"Invalid one-hot label for {case}: {one_hot}")
        grade = GRADES[one_hot.index(1)]
        source.append((case, grade)); strata[grade].append(case)
    if len(source) != 100 or len({x[0] for x in source}) != 100:
        raise SystemExit(f"Expected 100 unique cases, found {len(source)}")
    rng, splits = random.Random(args.seed), {}
    for grade in GRADES: splits.update(allocate(strata[grade], rng))
    rows = []
    for case, grade in sorted(source):
        case_dir = modalities / case
        cfp = args.raw_dir / "data/train" / f"{case}.jpg"
        mask = args.raw_dir / "mask_DC/train" / f"{case}.png"
        mhd = sorted(case_dir.glob("*_Sequence.mhd")); slice_dir = case_dir / case
        slices = sorted(slice_dir.glob("*_image.jpg")) if slice_dir.is_dir() else []
        if mhd:
            oct_path, oct_format, n_slices = mhd[0], "mhd_raw", ""
            if not mhd[0].with_suffix(".raw").is_file(): raise SystemExit(f"Missing raw payload: {case}")
        elif len(slices) == 256:
            oct_path, oct_format, n_slices = slice_dir, "jpg_slices", 256
        else: raise SystemExit(f"Invalid OCT for {case}: mhd={len(mhd)}, slices={len(slices)}")
        for kind, path in (("CFP", cfp), ("mask", mask)):
            if not path.is_file(): raise SystemExit(f"Missing {kind} for {case}: {path}")
        rows.append({"dataset":"gamma", "case_id":case, "patient_id":case,
            "split":splits[case], "label":int(grade != "non"), "grade":grade,
            "cfp_path":rel(cfp, root), "oct_path":rel(oct_path, root),
            "oct_format":oct_format, "oct_slices":n_slices,
            "disc_cup_mask_path":rel(mask, root)})
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    with args.manifest.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0])); w.writeheader(); w.writerows(rows)
    summary = {"seed":args.seed, "cases":len(rows),
        "grade_counts":dict(Counter(x["grade"] for x in rows)),
        "label_counts":dict(Counter(str(x["label"]) for x in rows)),
        "split_counts":dict(Counter(x["split"] for x in rows)),
        "split_grade_counts":{f"{s}|{g}":n for (s,g),n in sorted(Counter((x["split"],x["grade"]) for x in rows).items())},
        "oct_format_counts":dict(Counter(x["oct_format"] for x in rows)), "manifest":str(args.manifest)}
    args.summary.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n")
    print(json.dumps(summary, indent=2, sort_keys=True))

if __name__ == "__main__": main()
